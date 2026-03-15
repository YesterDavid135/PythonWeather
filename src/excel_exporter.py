from __future__ import annotations

import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.legend import Legend
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, RichTextProperties
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class ExcelExporter:

    # Farben
    HEADER_COLOR = "4472C4"
    HEADER_FONT_COLOR = "FFFFFF"
    TITLE_COLOR = "2F5496"
    ALT_ROW_COLOR = "D9E2F3"

    def __init__(self):
        self.header_font = Font(bold=True, color=self.HEADER_FONT_COLOR, size=11)
        self.header_fill = PatternFill(
            start_color=self.HEADER_COLOR,
            end_color=self.HEADER_COLOR,
            fill_type="solid",
        )
        self.title_font = Font(bold=True, color=self.TITLE_COLOR, size=16)
        self.subtitle_font = Font(bold=True, color=self.TITLE_COLOR, size=13)
        self.thin_side = Side(style="thin", color="B4C6E7")
        self.thin_border = Border(
            left=self.thin_side,
            right=self.thin_side,
            top=self.thin_side,
            bottom=self.thin_side,
        )
        self.alt_row_fill = PatternFill(
            start_color=self.ALT_ROW_COLOR,
            end_color=self.ALT_ROW_COLOR,
            fill_type="solid",
        )

    def export(
        self,
        city: str,
        current_data: dict,
        daily_data: list[dict],
        hourly_data: list[dict],
        output_dir: str = "./out",
    ) -> str:
        """Erstellt eine Excel-Datei mit Wetterdaten."""
        wb = Workbook()

        self._create_current_weather_sheet(wb, city, current_data)
        self._create_forecast_sheet(wb, city, daily_data, hourly_data)

        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
        filename = f"{output_dir}/Wetter_{city}_{timestamp}.xlsx"

        os.makedirs(output_dir, exist_ok=True)

        try:
            wb.save(filename)
        except PermissionError:
            raise PermissionError(
                f"Keine Schreibberechtigung für '{filename}'. "
                "Bitte Verzeichnis prüfen."
            )

        return filename

    #  Tabellenblatt 1: Aktuelles Wetter

    def _create_current_weather_sheet(
        self, wb: Workbook, city: str, data: dict
    ) -> None:
        ws = wb.active
        ws.title = "Aktuelles Wetter"

        # Titel
        ws.merge_cells("A1:B1")
        title_cell = ws["A1"]
        title_cell.value = f"Aktuelles Wetter für {city}"
        title_cell.font = self.title_font
        title_cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 30

        # Header
        row = 3
        for col, header in enumerate(["Eigenschaft", "Wert"], start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
            cell.alignment = Alignment(horizontal="center")

        # Datenzeilen
        for i, (key, value) in enumerate(data.items(), start=1):
            data_row = row + i
            key_cell = ws.cell(row=data_row, column=1, value=key)
            val_cell = ws.cell(row=data_row, column=2, value=value)

            key_cell.font = Font(bold=True)
            key_cell.border = self.thin_border
            val_cell.border = self.thin_border
            val_cell.alignment = Alignment(horizontal="center")

            # Abwechselnde Zeilenfarbe
            if i % 2 == 0:
                key_cell.fill = self.alt_row_fill
                val_cell.fill = self.alt_row_fill

        # Zeitstempel
        ts_row = row + len(data) + 2
        self._add_timestamp(ws, ts_row)

        # Spaltenbreite
        self._auto_column_width(ws)

    #  Tabellenblatt 2: Vorhersage

    def _create_forecast_sheet(
        self,
        wb: Workbook,
        city: str,
        daily_data: list[dict],
        hourly_data: list[dict],
    ) -> None:
        ws = wb.create_sheet(title="Vorhersage")

        # --- Titel ---
        ws.merge_cells("A1:H1")
        title_cell = ws["A1"]
        title_cell.value = f"Wettervorhersage für {city}"
        title_cell.font = self.title_font
        title_cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 30

        # --- Tagesvorhersage ---
        current_row = 3
        ws.cell(row=current_row, column=1, value="Tagesvorhersage").font = (
            self.subtitle_font
        )
        current_row += 1

        if daily_data:
            headers = list(daily_data[0].keys())
            daily_header_row = current_row
            self._write_header_row(ws, current_row, headers)
            current_row += 1

            daily_start_row = current_row
            for i, entry in enumerate(daily_data):
                for col, key in enumerate(headers, start=1):
                    cell = ws.cell(
                        row=current_row, column=col, value=entry.get(key, "")
                    )
                    cell.border = self.thin_border
                    cell.alignment = Alignment(horizontal="center")
                    if i % 2 == 0:
                        cell.fill = self.alt_row_fill
                current_row += 1
            daily_end_row = current_row - 1

            # Bedingte Formatierung auf Temperatur-Spalten (Spalten B, C, D)
            for col_letter in ["B", "C", "D"]:
                cell_range = f"{col_letter}{daily_start_row}:{col_letter}{daily_end_row}"
                ws.conditional_formatting.add(
                    cell_range,
                    ColorScaleRule(
                        start_type="num", start_value=-10, start_color="0000FF",
                        mid_type="num", mid_value=15, mid_color="FFFF00",
                        end_type="num", end_value=40, end_color="FF0000",
                     ),
                )

            # Temperatur-Liniendiagramm
            self._add_temperature_chart(
                ws, daily_header_row, daily_start_row, daily_end_row, len(daily_data)
            )

        # --- Lücke ---
        current_row += 2

        # --- Stündliche Vorhersage ---
        ws.cell(row=current_row, column=1, value="Stündliche Vorhersage").font = (
            self.subtitle_font
        )
        current_row += 1

        if hourly_data:
            headers = list(hourly_data[0].keys())
            hourly_header_row = current_row
            self._write_header_row(ws, current_row, headers)
            current_row += 1

            hourly_start_row = current_row
            for i, entry in enumerate(hourly_data):
                for col, key in enumerate(headers, start=1):
                    cell = ws.cell(
                        row=current_row, column=col, value=entry.get(key, "")
                    )
                    cell.border = self.thin_border
                    cell.alignment = Alignment(horizontal="center")
                    if i % 2 == 0:
                        cell.fill = self.alt_row_fill
                current_row += 1
            hourly_end_row = current_row - 1

            # Bedingte Formatierung auf Temperatur-Spalte (D) der stündlichen Daten
            temp_range = f"D{hourly_start_row}:D{hourly_end_row}"
            ws.conditional_formatting.add(
                temp_range,
                ColorScaleRule(
                    start_type="num", start_value=-10, start_color="0000FF",
                    mid_type="num", mid_value=15, mid_color="FFFF00",
                    end_type="num", end_value=40, end_color="FF0000",
                ),
            )

            # Helper-Spalte: Datum + Uhrzeit kombiniert für Diagramm-Labels
            label_col = len(headers) + 1
            for r in range(hourly_start_row, hourly_end_row + 1):
                date_val = ws.cell(row=r, column=1).value
                time_val = ws.cell(row=r, column=2).value
                ws.cell(row=r, column=label_col, value=f"{date_val} {time_val}")
            ws.column_dimensions[get_column_letter(label_col)].hidden = True


            # Temperatur & Niederschlag Kombi-Diagramm
            self._add_hourly_temp_precip_chart(
                ws, hourly_header_row, hourly_start_row, hourly_end_row,
                current_row + 2, label_col
            )

        # Zeitstempel
        current_row += 34
        self._add_timestamp(ws, current_row)

        # Spaltenbreite
        self._auto_column_width(ws)

    #  Diagramme

    def _add_temperature_chart(
        self,
        ws,
        header_row: int,
        start_row: int,
        end_row: int,
        num_days: int,
    ) -> None:
        """Fügt ein Liniendiagramm mit Temperaturverlauf hinzu."""
        chart = LineChart()
        chart.title = "Temperaturverlauf (Tagesvorhersage)"
        chart.y_axis.title = "Temperatur (°C)"
        chart.x_axis.title = "Datum"
        chart.style = 10
        chart.width = 20
        chart.height = 12

        # Datenreihen: Spalte B (Temperatur), C (Höchst), D (Tiefst)
        labels = Reference(ws, min_col=1, min_row=start_row, max_row=end_row)

        for col in [2, 3, 4]:
            data = Reference(
                ws, min_col=col, min_row=header_row, max_row=end_row
            )
            chart.add_data(data, titles_from_data=True)

        chart.set_categories(labels)

        # Legende anzeigen
        chart.legend = Legend()
        chart.legend.position = "b"

        # Farben für die Linien
        if len(chart.series) >= 3:
            chart.series[0].graphicalProperties.line.solidFill = "9BBB59"  # Durchschnitt: grün
            chart.series[1].graphicalProperties.line.solidFill = "C0504D"  # Höchst: rot
            chart.series[2].graphicalProperties.line.solidFill = "4F81BD"  # Tiefst: blau

        # Platzierung rechts neben der Tabelle
        chart_col = get_column_letter(len(list(ws.iter_cols())) + 6)
        ws.add_chart(chart, f"{chart_col}{header_row}")

    def _add_hourly_temp_precip_chart(
        self,
        ws,
        header_row: int,
        start_row: int,
        end_row: int,
        placement_row: int,
        label_col: int,
    ) -> None:
        """Fügt ein Kombi-Diagramm mit Temperatur (Linie, rot) und Niederschlag (Balken, blau) hinzu."""
        # Balkendiagramm für Niederschlag (Spalte I = 9)
        bar_chart = BarChart()
        bar_chart.type = "col"
        bar_chart.title = "Stündliche Temperatur & Niederschlag"
        bar_chart.y_axis.title = "Niederschlag (mm)"
        bar_chart.x_axis.title = "Datum / Uhrzeit"
        bar_chart.style = 10
        bar_chart.width = 24
        bar_chart.height = 14

        labels = Reference(ws, min_col=label_col, min_row=start_row, max_row=end_row)

        precip_data = Reference(ws, min_col=9, min_row=header_row, max_row=end_row)
        bar_chart.add_data(precip_data, titles_from_data=True)
        bar_chart.set_categories(labels)

        # Niederschlag-Balken blau
        if bar_chart.series:
            bar_chart.series[0].graphicalProperties.solidFill = "4F81BD"

        # Liniendiagramm für Temperatur (Spalte D = 4)
        line_chart = LineChart()
        line_chart.y_axis.title = "Temperatur (°C)"
        line_chart.y_axis.axId = 200

        temp_data = Reference(ws, min_col=4, min_row=header_row, max_row=end_row)
        line_chart.add_data(temp_data, titles_from_data=True)
        line_chart.set_categories(labels)

        # Temperatur-Linie rot
        if line_chart.series:
            line_chart.series[0].graphicalProperties.line.solidFill = "C0504D"
            line_chart.series[0].graphicalProperties.line.width = 22000

        # Legende anzeigen
        bar_chart.legend = Legend()
        bar_chart.legend.position = "b"

        # X-Achse: Uhrzeiten anzeigen mit Rotation
        bar_chart.x_axis.tickLblPos = "low"
        bar_chart.x_axis.txPr = RichText(
            p=[Paragraph(
                pPr=ParagraphProperties(
                    defRPr=CharacterProperties(sz=800)
                ),
                endParaRPr=CharacterProperties(sz=800),
            )],
            bodyPr=RichTextProperties(rot=-5400000),
        )

        # Zweite Y-Achse rechts für Temperatur
        line_chart.y_axis.crosses = "max"
        bar_chart += line_chart

        ws.add_chart(bar_chart, f"B{placement_row}")

    #  Hilfsmethoden

    def _write_header_row(self, ws, row: int, headers: list[str]) -> None:
        """Schreibt eine formatierte Header-Zeile."""
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
            cell.alignment = Alignment(horizontal="center")

    def _auto_column_width(self, ws) -> None:
        """Passt die Spaltenbreite automatisch an den Inhalt an."""
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 4, 40)

    def _add_timestamp(self, ws, row: int) -> None:
        """Fuegt einen Erstellungs-Zeitstempel hinzu."""
        timestamp = datetime.now().strftime("%d.%m.%Y %H:%M:%S")
        cell = ws.cell(row=row, column=1, value=f"Erstellt am: {timestamp}")
        cell.font = Font(italic=True, color="808080", size=9)
